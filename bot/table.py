import os
import openpyxl
from openpyxl.drawing.image import Image

def generate_track(client_code, products):
    wb = openpyxl.load_workbook("pattern/track.xlsx")
    sheet = wb.active

    sheet["C5"].value = client_code

    row_index = 10

    for product in products:
        photo_patch, trek_code, quantity = product

        quantity_cell = sheet[f'D{row_index}']
        quantity_cell.value = quantity

        image = Image(photo_patch)
        if image.width >= image.height:
            coef = image.height / image.width
            image.width = 180
            image.height = image.width * coef
        else:
            coef = image.width / image.height
            image.height = 180
            image.width = image.height * coef
        sheet.add_image(image, f'B{row_index}')

        trek_code_cell = sheet[f'C{row_index}']
        trek_code_cell.value = trek_code
        row_index += 1

    name = os.path.join("tables", f'{client_code}.xlsx')
    wb.save(name)
    return name




def generate_ransom(client_code, products):
    wb = openpyxl.load_workbook("pattern/ransom.xlsx")
    sheet = wb.active

    sheet.merge_cells('C4:D5')

    sheet["C4"].value = client_code

    row_index = 10

    for product in products:
        photo_patch, link, comment, quantity, price = product

        quantity_cell = sheet[f'E{row_index}']
        quantity_cell.value = quantity

        image = Image(photo_patch)
        if image.width >= image.height:
            coef = image.height / image.width
            image.width = 180
            image.height = image.width * coef
        else:
            coef = image.width / image.height
            image.height = 180
            image.width = image.height * coef
        sheet.add_image(image, f'B{row_index}')

        sheet[f"C{row_index}"] = link
        sheet[f"D{row_index}"] = comment
        sheet[f"F{row_index}"] = price

        row_index += 1

    name = os.path.join("tables", f'{client_code}.xlsx')
    wb.save(name)
    return name